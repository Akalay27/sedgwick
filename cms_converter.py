#!/usr/bin/env python3

# -----------------------------------------------------------------------------
# Made by Adam Kalayjian 2025
# This script converts PDFs to JPEGs using an Excel mapping file.
# It reads the mapping from an Excel file, processes each PDF, and saves
# the pages as JPEGs named with their corresponding XML IDs + page numbers.
# It also crops whitespace around the images.

# adam@kalayjian.org
# -----------------------------------------------------------------------------

import os
import sys
import logging
from glob import glob
import subprocess
from rich.prompt import Prompt   # NEW

import click
from openpyxl import load_workbook
from pdf2image import convert_from_path
from rich.console import Console
from rich.progress import (
    Progress,
    SpinnerColumn,
    TextColumn,
    BarColumn,
    MofNCompleteColumn,
    TimeElapsedColumn,
)
from rich.logging import RichHandler
from concurrent.futures import ThreadPoolExecutor, wait, as_completed

# -----------------------------------------------------------------------------
# Setup console and logging
# -----------------------------------------------------------------------------
console = Console()
logging.basicConfig(
    level=logging.INFO,
    format="%(message)s",
    datefmt="[%X]",
    handlers=[RichHandler(console=console, show_path=False, show_time=False)],
)
logger = logging.getLogger("rich")


# -----------------------------------------------------------------------------
# Core functions
# -----------------------------------------------------------------------------
def build_date_to_xml_map(
    excel_path: str,
    date_col: str,
    xml_id_col: str
) -> dict[str, str]:
    """
    Read all sheets in the workbook and return a dict mapping
    date_string -> xml_id
    """
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    mapping = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Find column indexes for "Date" and "XML ID" in the header row:
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        try:
            date_idx   = headers.index(date_col)
            xml_id_idx = headers.index(xml_id_col)
        except ValueError:
            # One of the columns is missing in this sheet
            logger.log(f"⚠️  Sheet '{sheet_name}' missing column '{date_col}' or '{xml_id_col}', skipping")
            continue

        # Iterate remaining rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            date_val = row[date_idx]
            xml_id   = row[xml_id_idx]
            if date_val is None or xml_id is None:
                continue
            date_str = str(date_val).strip()
            if date_str in mapping:
                logger.warning(
                    f"Duplicate date '{date_str}' in sheet '{sheet_name}', "
                    f"overwriting '{mapping[date_str]}' with '{xml_id}'"
                )
            mapping[date_str] = str(xml_id).strip()

    return mapping


def crop_whitespace(img, threshold: int):
    """
    Crop whitespace around an image by detecting borders where
    all pixels are above the threshold.
    """
    gray = img.convert('L')
    mask = gray.point(lambda x: 255 if x < threshold else 0, '1')
    bbox = mask.getbbox()
    return img.crop(bbox) if bbox else img

def process_pdf(
    pdf_path: str,
    date_map: dict[str, str],
    input_dir: str,
    output_dir: str,
    threshold: int,
    poppler_path: str | None,
    progress: Progress,
    pdf_task_id: int
):
    """
    Convert each page in pdf_path to JPEGs, cropping whitespace
    and naming them '{xml_id}-p{page_number}.jpg'.
    """
    base = os.path.splitext(os.path.basename(pdf_path))[0]
    xml_id = date_map.get(base)
    if not xml_id:
        logger.warning(f"No XML ID for '{base}', skipping")
        progress.update(pdf_task_id, advance=1)
        return

    try:
        pages = convert_from_path(pdf_path, poppler_path=poppler_path)
    except Exception as e:
        logger.warning(f"Failed to convert '{pdf_path}': {e}")
        progress.update(pdf_task_id, advance=1)
        return
    
    rel_path = os.path.relpath(os.path.dirname(pdf_path), input_dir)
    sub_output_dir = os.path.join(output_dir, rel_path)
    os.makedirs(sub_output_dir, exist_ok=True)
    
    saved_count = 0
    for i, img in enumerate(pages, start=1):
        cropped = crop_whitespace(img, threshold)
        if cropped.size != img.size:
            logger.debug(f"– cropped whitespace on page {i}")
        out_name = f"{xml_id}-p{i}.jpg"
        out_path = os.path.join(sub_output_dir, out_name)
        try:
            cropped.save(out_path, "JPEG")
            saved_count += 1
        except Exception as e:
            logger.warning(f"Failed to save '{out_name}': {e}")
    if saved_count == len(pages):
        logger.info(f"Converted '{base}' to {saved_count} images")
    else:
        logger.warning(f"Only {saved_count} out of {len(pages)} images saved for '{base}'")
    return saved_count


def _open_dialog(select_dir: bool) -> str | None:
    """Open a native file/folder chooser and return the selected path, or None."""
    if sys.platform == "darwin":                      # macOS ➜ AppleScript
        script = 'choose folder with prompt "Select a folder"' if select_dir \
                 else 'choose file with prompt "Select a file"'
        cmd = ["osascript", "-e", f'POSIX path of ({script})']
        result = subprocess.run(cmd, capture_output=True, text=True)
        return result.stdout.strip() if result.returncode == 0 else None
    if sys.platform.startswith("linux"):              # Linux ➜ zenity
        zen_cmd = ["zenity", "--file-selection", "--title=Select a folder" if select_dir else "--title=Select a file"]
        if select_dir:
            zen_cmd.append("--directory")
        result = subprocess.run(zen_cmd, capture_output=True, text=True)
        return result.stdout.strip() if result.returncode == 0 else None

    if sys.platform.startswith("win"):                # Windows ➜ PowerShell
        ps = r'''
        Add-Type -AssemblyName System.Windows.Forms
        $d = New-Object System.Windows.Forms.{0}Dialog
        if ($d.ShowDialog() -eq "OK") {{ Write-Output $d.SelectedPath }}
        '''.format("FolderBrowser" if select_dir else "OpenFile")
        result = subprocess.run(["powershell", "-NoProfile", ps],
                                capture_output=True, text=True)
        return result.stdout.strip() if result.returncode == 0 else None
    return None

def _ask_path(label: str, default: str, select_dir: bool) -> str:
    """
    Prompt user: Enter → default, 'c' → open chooser, anything else → manual entry.
    """
    prompt_text = f"[bold]{label}[/] (Enter=‘{default}’, [cyan]c[/]=choose, other=type path)"
    answer = Prompt.ask(prompt_text, default="").strip()

    # Enter → keep default
    if answer == "":
        # logger.info(f"{label}: {default}")
        return default

    # 'c' → dialog
    if answer.lower() == "c":
        chosen = _open_dialog(select_dir)
        if chosen:
            logger.info(f"{label}: {chosen}")
            return chosen
        logger.warning("Could not open a file‑chooser; using default.")
        # logger.info(f"{label}: {default}")
        return default

    # typed path
    # logger.info(f"{label}: {answer}")
    return answer

# -----------------------------------------------------------------------------
# CLI entrypoint
# -----------------------------------------------------------------------------
@click.command(context_settings={"help_option_names": ["-h", "--help"]})
@click.option("--input-dir",    "-i", default="input",      show_default=True,
              help="Directory containing PDF files")
@click.option("--output-dir",   "-o", default="output",     show_default=True,
              help="Where to write JPEGs")
@click.option("--spreadsheet",  "-s", default="Workflow Inventory 1796-1867.xlsx",
              show_default=True, help="Excel file mapping dates → XML IDs")
@click.option("--date-col",     default="Date",              show_default=True,
              help="Column name for dates in spreadsheet")
@click.option("--xml-id-col",   default="XML ID",            show_default=True,
              help="Column name for XML IDs in spreadsheet")
@click.option("--threshold",    "-t", default=250,           show_default=True,
              help="Whitespace‐detection threshold (0–255)")
@click.option("--poppler-path", "-p", default=None,
              help="Path to Poppler binaries (if not in $PATH)")
@click.option("--verbose/--quiet", default=False,
              help="Enable DEBUG logging when set; otherwise INFO")
@click.option("--prompt/--no-prompt", default=True,
              help="Prompt for paths instead of using defaults")
def cli(
    input_dir: str,
    output_dir: str,
    spreadsheet: str,
    date_col: str,
    xml_id_col: str,
    threshold: int,
    poppler_path: str,
    verbose: bool,
    prompt: bool
):  

    if prompt:
        input_dir  = _ask_path("Input directory",  input_dir,  select_dir=True)
        output_dir = _ask_path("Output directory", output_dir, select_dir=True)
        spreadsheet = _ask_path("Spreadsheet file", spreadsheet, select_dir=False)
    
    # adjust logging level
    logger.setLevel(logging.DEBUG if verbose else logging.INFO)

    if not os.path.isdir(input_dir):
        logger.error(f"Input directory '{input_dir}' not found.")
        sys.exit(1)
    os.makedirs(output_dir, exist_ok=True)

    logger.info(f"Reading mapping from '{spreadsheet}'…")
    date_map = build_date_to_xml_map(spreadsheet, date_col, xml_id_col)
    logger.info(f"→ {len(date_map)} entries loaded")

    pdfs = []
    for root, _, files in os.walk(input_dir):
        for file in files:
            if file.lower().endswith(".pdf"):
                pdfs.append(os.path.join(root, file))
    if not pdfs:
        logger.warning(f"No PDFs in '{input_dir}'.")
        return

    # Single Rich Progress for both PDF files and pages
    with Progress(
        SpinnerColumn("dots", style="green", speed=0.5),
        TextColumn("[bold magenta]{task.description}"),
        BarColumn(bar_width=None),
        MofNCompleteColumn(),
        TimeElapsedColumn(),
        console=console,
        transient=True,
        expand=True
    ) as progress:
        pdf_task_id = progress.add_task("Converting PDFs...", total=len(pdfs))
        with ThreadPoolExecutor(max_workers=os.cpu_count()) as pool:
            futures = [
                pool.submit(
                    process_pdf,
                    pdf,
                    date_map,
                    input_dir,
                    output_dir,
                    threshold,
                    poppler_path,
                    progress,
                    pdf_task_id
                )
                for pdf in pdfs
            ]
            
            total_pages = 0
            for future in as_completed(futures):
                result = future.result()
                if result:
                    total_pages += result
                progress.update(pdf_task_id, advance=1)
    logger.info("Conversion complete!")

    logger.info(f"Created {total_pages} images in '{output_dir}'")
if __name__ == "__main__":
    cli()
